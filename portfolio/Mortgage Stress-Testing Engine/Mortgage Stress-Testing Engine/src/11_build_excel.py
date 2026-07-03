"""
11_build_excel.py — Build the Excel deliverable (Stress_Test_Results.xlsx).

Four sheets: Summary (KPI cards via cross-sheet formulas), Projection (13-quarter
loss paths + line chart), Segments (loss by FICO/LTV band with conditional
formatting), and a formula-driven Stress Calculator (change unemployment / HPI
inputs -> projected loss recomputes live from the fitted hazard coefficients).

Follows the xlsx skill: professional font, blue inputs / black formulas / green
cross-sheet links, currency/percent number formats, formulas (not hardcoded
calcs). Run scripts/recalc.py afterwards to populate values + check errors.
"""
from __future__ import annotations
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from config import OUTPUTS, DASHBOARD, ROOT

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")      # inputs
BLACK = Font(name=FONT, color="000000")     # formulas
GREEN = Font(name=FONT, color="008000")     # cross-sheet links
H1 = Font(name=FONT, size=16, bold=True, color="1F2937")
H2 = Font(name=FONT, size=12, bold=True, color="FFFFFF")
LBL = Font(name=FONT, size=9, color="6B7280")
HDR_FILL = PatternFill("solid", fgColor="1E293B")
YEL = PatternFill("solid", fgColor="FFFF00")
CARD = PatternFill("solid", fgColor="EEF2FF")
thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H2; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")


def main():
    proj = pd.read_csv(OUTPUTS / "expected_loss_by_scenario.csv")
    seg = pd.read_csv(OUTPUTS / "loss_by_segment.csv")
    sim = json.loads(open(DASHBOARD / "data.js").read()
                     .replace("window.DASH_DATA = ", "").rstrip(";"))["simulator"]

    wb = Workbook()

    # ================= SUMMARY =================
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Severely Adverse — Mortgage Stress-Test Results"; ws["A1"].font = H1
    ws["A2"] = ("Loan-level PD model on real Freddie Mac mortgages, projected under the "
                "Federal Reserve's 2026 baseline vs severely-adverse (CCAR/DFAST) scenarios.")
    ws["A2"].font = LBL
    cards = [("Portfolio (UPB)", "='Projection'!$H$2", '$#,##0,,"M"'),
             ("Loans", "=Projection!$I$2", "#,##0"),
             ("Baseline loss rate", "=Projection!B15", "0.00%"),
             ("Severely-adverse loss", "=Projection!C15", "0.00%"),
             ("Severe ÷ baseline", "=Projection!C15/Projection!B15", '0.0"x"')]
    r = 4
    for i, (lbl, f, fmt) in enumerate(cards):
        col = 1 + (i % 3) * 3
        rr = r + (i // 3) * 3
        lc = ws.cell(row=rr, column=col, value=lbl); lc.font = LBL
        vc = ws.cell(row=rr + 1, column=col, value=f)
        vc.font = Font(name=FONT, size=18, bold=True, color="1D4ED8"); vc.number_format = fmt
        vc.fill = CARD; vc.border = BORDER
    ws["A11"] = "Headline"; ws["A11"].font = Font(name=FONT, bold=True)
    ws["A12"] = ('=CONCATENATE("Projected 13-quarter loss = ",TEXT(Projection!B15,"0.00%"),'
                 '" (baseline) vs ",TEXT(Projection!C15,"0.00%")," (severely adverse) — ",'
                 'TEXT(Projection!C15/Projection!B15,"0.0"),"x baseline.")')
    ws["A12"].font = Font(name=FONT, italic=True)
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 14

    # ================= PROJECTION =================
    wp = wb.create_sheet("Projection")
    wp.sheet_view.showGridLines = False
    wp["A1"] = "Quarter"; wp["B1"] = "Baseline cum loss %"; wp["C1"] = "Severe cum loss %"
    wp["D1"] = "Baseline cum default %"; wp["E1"] = "Severe cum default %"
    wp["F1"] = "Unemployment (severe) %"; wp["H1"] = "Portfolio UPB"; wp["I1"] = "Loans"
    style_header(wp, 1, 6)
    b = proj[proj.scenario == "baseline"].reset_index(drop=True)
    s = proj[proj.scenario == "severely_adverse"].reset_index(drop=True)
    for i in range(len(b)):
        rr = i + 2
        wp.cell(row=rr, column=1, value=b.loc[i, "quarter"]).font = BLACK
        wp.cell(row=rr, column=2, value=b.loc[i, "cumulative_loss_rate_pct"] / 100).number_format = "0.00%"
        wp.cell(row=rr, column=3, value=s.loc[i, "cumulative_loss_rate_pct"] / 100).number_format = "0.00%"
        wp.cell(row=rr, column=4, value=b.loc[i, "cumulative_default_rate_pct"] / 100).number_format = "0.00%"
        wp.cell(row=rr, column=5, value=s.loc[i, "cumulative_default_rate_pct"] / 100).number_format = "0.00%"
        wp.cell(row=rr, column=6, value=s.loc[i, "unemployment_rate"] / 100).number_format = "0.0%"
    wp["H2"] = float(proj["portfolio_upb0"].iloc[0]); wp["H2"].number_format = "$#,##0"
    wp["I2"] = int(proj["n_loans"].iloc[0]); wp["I2"].number_format = "#,##0"
    # summary formulas (final-quarter values)
    last = len(b) + 1
    wp["A15"] = "Final"; wp["A15"].font = Font(name=FONT, bold=True)
    wp["B15"] = f"=B{last}"; wp["C15"] = f"=C{last}"
    for c in "BC": wp[f"{c}15"].number_format = "0.00%"; wp[f"{c}15"].font = Font(name=FONT, bold=True)

    chart = LineChart(); chart.title = "Cumulative portfolio loss — Fed 2026 scenarios"
    chart.y_axis.title = "Cumulative loss %"; chart.height = 8; chart.width = 16
    data = Reference(wp, min_col=2, max_col=3, min_row=1, max_row=len(b) + 1)
    cats = Reference(wp, min_col=1, min_row=2, max_row=len(b) + 1)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    wp.add_chart(chart, "K1")
    wp.column_dimensions["A"].width = 10
    for col in "BCDEF": wp.column_dimensions[col].width = 18
    wp.column_dimensions["H"].width = 16

    # ================= SEGMENTS =================
    wsg = wb.create_sheet("Segments")
    wsg.sheet_view.showGridLines = False
    wsg["A1"] = "Loss rate by risk segment (severely adverse)"; wsg["A1"].font = H1
    row = 3
    for dim, title in [("fico_band", "FICO band"), ("ltv_band", "Original LTV band")]:
        d = seg[(seg.dimension == dim) & (seg.scenario == "severely_adverse")].copy()
        d = d.sort_values("segment")
        wsg.cell(row=row, column=1, value=title); wsg.cell(row=row, column=2, value="Loans")
        wsg.cell(row=row, column=3, value="Loss rate %"); wsg.cell(row=row, column=4, value="Expected loss")
        style_header(wsg, row, 4)
        start = row + 1
        for _, x in d.iterrows():
            row += 1
            wsg.cell(row=row, column=1, value=x["segment"]).font = BLACK
            wsg.cell(row=row, column=2, value=int(x["loans"])).number_format = "#,##0"
            lc = wsg.cell(row=row, column=3, value=x["loss_rate_pct"] / 100); lc.number_format = "0.00%"
            wsg.cell(row=row, column=4, value=x["exp_loss"]).number_format = "$#,##0"
        wsg.conditional_formatting.add(
            f"C{start}:C{row}",
            ColorScaleRule(start_type="min", start_color="FEE2E2",
                           end_type="max", end_color="DC2626"))
        row += 2
    for col, w in [("A", 16), ("B", 12), ("C", 12), ("D", 16)]:
        wsg.column_dimensions[col].width = w

    # ================= STRESS CALCULATOR =================
    wc = wb.create_sheet("Stress Calculator")
    wc.sheet_view.showGridLines = False
    wc["A1"] = "Interactive Stress Calculator (change the blue inputs)"; wc["A1"].font = H1
    wc["A2"] = ("Illustrative — driven by the fitted discrete-time hazard on a representative "
                "loan. Change B4/B5; the loss path recomputes."); wc["A2"].font = LBL
    wc["A4"] = "Unemployment peak (%)"; wc["B4"] = 10.0
    wc["A5"] = "House-price decline (%)"; wc["B5"] = 30.0
    for cell in ("B4", "B5"):
        wc[cell].font = BLUE; wc[cell].fill = YEL; wc[cell].border = BORDER; wc[cell].number_format = "0.0"

    # assumptions: feature | coef | mean | scale  (rows 8-14)
    feats = sim["features"]
    wc["E7"] = "Feature"; wc["F7"] = "Coef"; wc["G7"] = "Mean"; wc["H7"] = "Scale"
    style_header(wc, 7, 4)
    frow = {}
    for i, f in enumerate(feats):
        rr = 8 + i; frow[f] = rr
        wc.cell(row=rr, column=5, value=f).font = BLACK
        wc.cell(row=rr, column=6, value=sim["coef"][f]).font = BLACK
        wc.cell(row=rr, column=7, value=sim["scaler_mean"][f]).font = BLACK
        wc.cell(row=rr, column=8, value=sim["scaler_scale"][f]).font = BLACK
    # scalar assumptions in K/L
    scal = [("intercept", sim["intercept"]), ("fico_rep", sim["rep_loan"]["fico"]),
            ("oltv_rep", sim["rep_loan"]["oltv"]), ("dti_rep", sim["rep_loan"]["odti"]),
            ("rate_rep", sim["rep_loan"]["orig_rate"]), ("lgd", sim["lgd"]),
            ("prepay_q", sim["prepay_hazard_q"]), ("base_unemp", sim["base_unemployment"]),
            ("start_age", sim["start_age_q"]),
            ("portfolio_upb", float(proj["portfolio_upb0"].iloc[0]))]
    srow = {}
    for i, (k, v) in enumerate(scal):
        rr = 8 + i; srow[k] = rr
        wc.cell(row=rr, column=11, value=k).font = LBL
        wc.cell(row=rr, column=12, value=v).font = BLACK
    def SC(k): return f"$L${srow[k]}"          # scalar cell
    def CO(f): return f"$F${frow[f]}"
    def ME(f): return f"$G${frow[f]}"
    def SD(f): return f"$H${frow[f]}"

    # projection table rows 20..32 (t=1..13)
    hdr = ["t", "Age(q)", "Age^2", "Unemp%", "HPI idx", "Cur LTV", "z", "h_def",
           "Survival", "Exp def", "Cum def", "Cum loss %"]
    for j, h in enumerate(hdr):
        wc.cell(row=19, column=1 + j, value=h)
    style_header(wc, 19, len(hdr))
    first = 20
    for t in range(1, 14):
        rr = first + t - 1
        A = f"A{rr}"; B = f"B{rr}"; Cc = f"C{rr}"; Dd = f"D{rr}"; E = f"E{rr}"
        F = f"F{rr}"; G = f"G{rr}"; H = f"H{rr}"; I = f"I{rr}"; J = f"J{rr}"
        K = f"K{rr}"; L = f"L{rr}"
        wc[A] = t
        wc[B] = f"={SC('start_age')}+{A}"
        wc[Cc] = f"={B}^2"
        wc[Dd] = f"={SC('base_unemp')}+($B$4-{SC('base_unemp')})*MIN({A},7)/7"
        wc[E] = f"=100*(1-($B$5/100)*MIN({A},8)/8)"
        wc[F] = f"=MIN({SC('oltv_rep')}*100/{E},250)"
        wc[G] = (f"={SC('intercept')}"
                 f"+{CO('loan_age_q')}*(({B}-{ME('loan_age_q')})/{SD('loan_age_q')})"
                 f"+{CO('loan_age_q_sq')}*(({Cc}-{ME('loan_age_q_sq')})/{SD('loan_age_q_sq')})"
                 f"+{CO('fico')}*(({SC('fico_rep')}-{ME('fico')})/{SD('fico')})"
                 f"+{CO('current_ltv')}*(({F}-{ME('current_ltv')})/{SD('current_ltv')})"
                 f"+{CO('odti')}*(({SC('dti_rep')}-{ME('odti')})/{SD('odti')})"
                 f"+{CO('orig_rate')}*(({SC('rate_rep')}-{ME('orig_rate')})/{SD('orig_rate')})"
                 f"+{CO('unemployment_rate')}*(({Dd}-{ME('unemployment_rate')})/{SD('unemployment_rate')})")
        wc[H] = f"=1/(1+EXP(-{G}))"
        if t == 1:
            wc[I] = "=1"
            wc[K] = f"={J}"
        else:
            pr = rr - 1
            wc[I] = f"=I{pr}*(1-H{pr}-{SC('prepay_q')})"
            wc[K] = f"=K{pr}+{J}"
        wc[J] = f"={I}*{H}"
        wc[L] = f"={K}*{SC('lgd')}"
        wc[L].number_format = "0.00%"
        for col in "BCDEFG": wc[f"{col}{rr}"].number_format = "0.00"
    lastr = first + 12
    wc["A34"] = "Projected cumulative default"; wc["B34"] = f"=K{lastr}"; wc["B34"].number_format = "0.00%"
    wc["A35"] = "Projected loss rate"; wc["B35"] = f"=L{lastr}"; wc["B35"].number_format = "0.00%"
    wc["A36"] = "Projected expected loss ($)"; wc["B36"] = f"=B35*{SC('portfolio_upb')}"
    wc["B36"].number_format = "$#,##0"
    for a in ("A34", "A35", "A36"): wc[a].font = Font(name=FONT, bold=True)
    for a in ("B34", "B35", "B36"): wc[a].font = Font(name=FONT, bold=True, color="B91C1C")
    cchart = LineChart(); cchart.title = "Projected cumulative loss %"; cchart.height = 7; cchart.width = 14
    cdata = Reference(wc, min_col=12, min_row=19, max_row=lastr)
    ccats = Reference(wc, min_col=1, min_row=first, max_row=lastr)
    cchart.add_data(cdata, titles_from_data=True); cchart.set_categories(ccats)
    wc.add_chart(cchart, "E17")
    wc.column_dimensions["A"].width = 24
    for col in "BCDEFGHIJKL": wc.column_dimensions[col].width = 11

    out = ROOT / "excel" / "Stress_Test_Results.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
